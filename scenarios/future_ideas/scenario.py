"""
HELM Scenario: Future Research Idea Generation Benchmark

Paper: Can Large Language Models Unlock Novel Scientific Research Ideas?
       EMNLP 2025
       https://arxiv.org/abs/2409.06185
Code:  https://github.com/sandeep82945/Future-Idea-Generation

Task: Given a research paper (with the future work section removed), generate
future research ideas that could extend the paper's contributions. Tests
scientific creativity — the ability to identify novel, relevant, and feasible
research directions based on understanding of existing work.

Dataset: 458 papers across 5 scientific domains, stored as xlsx files in
  data/annotations/RealF/ on GitHub. Each row has the paper text without
  future work (full_text_WF) and the author-written future work section
  (Future_work) as the gold reference.

Domains (via domain= parameter):
  chemistry — 72 papers
  computer  — 88 papers
  economics — 61 papers
  medical   — 110 papers
  physics   — 127 papers

The full_text_WF field has the future work section deliberately removed so
the model must infer what directions are needed. The Future_work field
contains the authors' actual future work text as the gold reference.

Paper content is truncated to MAX_WORDS to stay within model context limits.
The future work section typically appears at the end of papers, so truncating
from the start preserves the bulk of the paper's technical content.

Prompt format (no explicit prompt in paper — standard instruction used):
  The following is the content of a research paper (future work section removed).
  Based on the paper's contributions and limitations, propose future research
  directions and ideas.

  {paper_text}

  Future research ideas:

Prompt source: Standard instruction format (paper does not publish exact prompts)
Fields used: full_text_WF (paper without future work, as input),
  Future_work (author-written future work section, as gold reference)
Fields skipped: full_text (includes future work — data leakage),
  Response_Chat (LLM-generated outputs, not ground truth),
  paper_id (metadata only)

Evaluation: open_ended (BLEU, ROUGE against author-written future work)
  For idea-level alignment scoring (IAScore, IDI), see metric_notes.md
"""

import io
import os
import urllib.request
from typing import List

import openpyxl

from helm.benchmark.scenarios.scenario import (
    Scenario, Instance, Input, Output, Reference,
    CORRECT_TAG, TEST_SPLIT,
)

_BASE_URL = (
    "https://raw.githubusercontent.com/sandeep82945/Future-Idea-Generation"
    "/main/data/annotations/RealF"
)

_DOMAIN_FILES = {
    "chemistry": "Idea_chemistry.xlsx",
    "computer": "idea_computer.xlsx",
    "economics": "idea_economics.xlsx",
    "medical": "idea_medical.xlsx",
    "physics": "idea_physics.xlsx",
}

# Truncate paper text to this many words to stay within model context.
# The future work section is typically at the end, so leading truncation
# preserves the paper's core technical content.
_MAX_WORDS = 2000


class FutureIdeasScenario(Scenario):
    """
    Future research idea generation from scientific papers.

    Given a paper (future work section removed), generate future research
    directions. Gold reference is the author's own future work section.

    domain= one of: chemistry, computer, economics, medical, physics
    """

    name = "future_ideas"
    description = "sandeep82945/Future-Idea-Generation"
    tags = ["creativity", "scientific_creativity", "idea_generation", "open_ended"]

    DOMAINS = list(_DOMAIN_FILES.keys())

    def __init__(self, domain: str):
        super().__init__()
        assert domain in self.DOMAINS, (
            f"domain must be one of {self.DOMAINS}, got '{domain}'"
        )
        self.domain = domain

    def get_instances(self, output_path: str) -> List[Instance]:
        os.makedirs(output_path, exist_ok=True)
        fname = _DOMAIN_FILES[self.domain]
        local_path = os.path.join(output_path, fname)

        if not os.path.exists(local_path):
            url = f"{_BASE_URL}/{fname}"
            urllib.request.urlretrieve(url, local_path)

        with open(local_path, "rb") as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()))
        ws = wb.active

        # Map column names to indices
        headers = [cell.value for cell in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}

        instances = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            full_text_wf = row[col["full_text_WF"] - 1]
            future_work = row[col["Future_work"] - 1]

            if not full_text_wf or not future_work:
                continue

            paper_text = str(full_text_wf).strip()
            gold_text = str(future_work).strip()

            # Truncate to avoid excessively long prompts
            words = paper_text.split()
            if len(words) > _MAX_WORDS:
                paper_text = " ".join(words[:_MAX_WORDS])

            prompt = (
                "The following is the content of a research paper "
                "(future work section removed). Based on the paper's "
                "contributions and limitations, propose future research "
                "directions and ideas.\n\n"
                f"{paper_text}\n\n"
                "Future research ideas:"
            )

            instances.append(Instance(
                input=Input(text=prompt),
                references=[Reference(Output(text=gold_text), tags=[CORRECT_TAG])],
                split=TEST_SPLIT,
            ))

        return instances
